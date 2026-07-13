import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math

# Buat workbook
wb = Workbook()

# ============================================
# SHEET 1: KORPUS & TOKENISASI
# ============================================
ws1 = wb.active
ws1.title = "1. Korpus & Tokenisasi"

# Header
ws1['A1'] = "DOKUMEN"
ws1['B1'] = "TEKS ASLI"
ws1['C1'] = "TEKS SETEMAH STEMMING (Sastrawi)"
ws1['D1'] = "KETERANGAN"

# Styling header
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Data
data_korpus = [
    ["D1", "web developer", "web develop", "Stemming: developer → develop"],
    ["D2", "web analyst", "web analyst", "Tidak berubah"],
    ["D3", "web developer analyst", "web develop analyst", "Stemming: developer → develop"]
]

for i, row in enumerate(data_korpus, start=2):
    for j, val in enumerate(row, start=1):
        ws1.cell(row=i, column=j, value=val)

# Catatan penting
ws1['A6'] = "CATATAN PENTING:"
ws1['A6'].font = Font(bold=True, color="FF0000")
ws1['A7'] = "1. Korpus ini dirancang agar setiap kata muncul minimal 2x (kompatibel dengan min_df=2)"
ws1['A8'] = "2. Stemming Sastrawi mengubah 'developer' menjadi 'develop'"
ws1['A9'] = "3. Total dokumen N = 3"

# Set column width
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 30
ws1.column_dimensions['D'].width = 40

# ============================================
# SHEET 2: MATRIKS TF & SUBLINEAR TF
# ============================================
ws2 = wb.create_sheet("2. Matriks TF")

# Matriks TF Mentah
ws2['A1'] = "MATRIKS TF MENTAH (Frekuensi Term)"
ws2['A1'].font = Font(bold=True, size=12)

ws2['A2'] = "Dokumen"
ws2['B2'] = "web"
ws2['C2'] = "develop"
ws2['D2'] = "analyst"

for cell in ws2[2]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Data TF
tf_data = [
    ["D1", 1, 1, 0],
    ["D2", 1, 0, 1],
    ["D3", 1, 1, 1]
]

for i, row in enumerate(tf_data, start=3):
    for j, val in enumerate(row, start=1):
        ws2.cell(row=i, column=j, value=val)

# Sublinear TF
ws2['A7'] = "SUBLINEAR TF (1 + ln(tf))"
ws2['A7'].font = Font(bold=True, size=12)

ws2['A8'] = "Dokumen"
ws2['B8'] = "web"
ws2['C8'] = "develop"
ws2['D8'] = "analyst"

for cell in ws2[8]:
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Hitung Sublinear TF
sublinear_data = [
    ["D1", 1 + math.log(1), 1 + math.log(1), 0],
    ["D2", 1 + math.log(1), 0, 1 + math.log(1)],
    ["D3", 1 + math.log(1), 1 + math.log(1), 1 + math.log(1)]
]

for i, row in enumerate(sublinear_data, start=9):
    for j, val in enumerate(row, start=1):
        if j == 1:
            ws2.cell(row=i, column=j, value=val)
        else:
            ws2.cell(row=i, column=j, value=round(val, 4))

ws2['A13'] = "CATATAN:"
ws2['A13'].font = Font(bold=True)
ws2['A14'] = "Karena semua TF bernilai 0 atau 1, dan ln(1) = 0, maka Sublinear TF = TF mentah"

# Set column width
for col in ['A', 'B', 'C', 'D']:
    ws2.column_dimensions[col].width = 15

# ============================================
# SHEET 3: PERHITUNGAN DF & SMOOTH IDF
# ============================================
ws3 = wb.create_sheet("3. Smooth IDF")

ws3['A1'] = "PERHITUNGAN DOCUMENT FREQUENCY (DF) & SMOOTH IDF"
ws3['A1'].font = Font(bold=True, size=12)

ws3['A3'] = "Kata (Vocabulary)"
ws3['B3'] = "DF (Jumlah Dokumen)"
ws3['C3'] = "Rumus Smooth IDF Scikit-Learn"
ws3['D3'] = "Nilai IDF"

for cell in ws3[3]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Data DF dan IDF
N = 3
df_data = [
    ["web", 3],
    ["develop", 2],
    ["analyst", 2]
]

for i, (word, df) in enumerate(df_data, start=4):
    ws3.cell(row=i, column=1, value=word)
    ws3.cell(row=i, column=2, value=df)
    
    # Rumus: ln((1+N)/(1+df)) + 1
    rumus = f"=LN((1+{N})/(1+{df}))+1"
    ws3.cell(row=i, column=3, value=rumus)
    
    # Hitung nilai
    idf_value = math.log((1 + N) / (1 + df)) + 1
    ws3.cell(row=i, column=4, value=round(idf_value, 4))
    ws3.cell(row=i, column=4).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

ws3['A8'] = "CATATAN:"
ws3['A8'].font = Font(bold=True)
ws3['A9'] = f"Total Dokumen N = {N}"
ws3['A10'] = "Rumus Smooth IDF: ln((1+N)/(1+df)) + 1"
ws3['A11'] = "Smooth IDF mencegah pembagian dengan nol dan menstabilkan bobot untuk term langka"

# Highlight contoh perhitungan
ws3['A13'] = "CONTOH PERHITUNGAN MANUAL untuk kata 'web':"
ws3['A13'].font = Font(bold=True, color="FF0000")
ws3['A14'] = f"IDF = ln((1+3)/(1+3)) + 1 = ln(4/4) + 1 = ln(1) + 1 = 0 + 1 = 1.0000"
ws3['A15'] = f"IDF = ln((1+3)/(1+2)) + 1 = ln(4/3) + 1 = 0.2877 + 1 = 1.2877"

for col in ['A', 'B', 'C', 'D']:
    ws3.column_dimensions[col].width = 30

# ============================================
# SHEET 4: MATRIKS TF-IDF & NORMALISASI L2
# ============================================
ws4 = wb.create_sheet("4. TF-IDF & L2 Norm")

ws4['A1'] = "MATRIKS TF-IDF MENTAH (Sublinear TF × IDF)"
ws4['A1'].font = Font(bold=True, size=12)

ws4['A3'] = "Dokumen"
ws4['B3'] = "web"
ws4['C3'] = "develop"
ws4['D3'] = "analyst"

for cell in ws4[3]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Hitung TF-IDF mentah
idf_values = {
    "web": 1.0000,
    "develop": 1.2877,
    "analyst": 1.2877
}

tfidf_data = [
    ["D1", 1 * idf_values["web"], 1 * idf_values["develop"], 0],
    ["D2", 1 * idf_values["web"], 0, 1 * idf_values["analyst"]],
    ["D3", 1 * idf_values["web"], 1 * idf_values["develop"], 1 * idf_values["analyst"]]
]

for i, row in enumerate(tfidf_data, start=4):
    for j, val in enumerate(row, start=1):
        if j == 1:
            ws4.cell(row=i, column=j, value=val)
        else:
            ws4.cell(row=i, column=j, value=round(val, 4))

# L2 Norm
ws4['A8'] = "L2 NORM (Panjang Vektor per Dokumen)"
ws4['A8'].font = Font(bold=True, size=12)

ws4['A10'] = "Dokumen"
ws4['B10'] = "L2 Norm = √(x₁² + x₂² + x₃²)"

for cell in ws4[10]:
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Hitung L2 Norm
l2_norms = []
for row in tfidf_data:
    values = [v for v in row[1:] if isinstance(v, (int, float))]
    l2 = math.sqrt(sum(v**2 for v in values))
    l2_norms.append(l2)
    ws4.cell(row=11 + tfidf_data.index(row), column=1, value=row[0])
    ws4.cell(row=11 + tfidf_data.index(row), column=2, value=round(l2, 4))
    ws4.cell(row=11 + tfidf_data.index(row), column=2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Vektor Fitur Akhir (Ternormalisasi L2)
ws4['A15'] = "VEKTOR FITUR AKHIR (TF-IDF Ternormalisasi L2)"
ws4['A15'].font = Font(bold=True, size=12, color="FF0000")

ws4['A17'] = "Dokumen"
ws4['B17'] = "web"
ws4['C17'] = "develop"
ws4['D17'] = "analyst"

for cell in ws4[17]:
    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")

# Hitung vektor ternormalisasi
for i, (row, l2) in enumerate(zip(tfidf_data, l2_norms), start=18):
    ws4.cell(row=i, column=1, value=row[0])
    for j, val in enumerate(row[1:], start=2):
        if isinstance(val, (int, float)) and val > 0:
            normalized = val / l2
            ws4.cell(row=i, column=j, value=round(normalized, 4))
        else:
            ws4.cell(row=i, column=j, value=0)

ws4['A22'] = "CATATAN:"
ws4['A22'].font = Font(bold=True)
ws4['A23'] = "Normalisasi L2 memastikan panjang vektor = 1 (standar Scikit-Learn)"
ws4['A24'] = "Vektor ini yang akan dikalikan dengan bobot koefisien model Logistic Regression"

for col in ['A', 'B', 'C', 'D']:
    ws4.column_dimensions[col].width = 20

# ============================================
# SHEET 5: SIMULASI INFERENSI LOGISTIC REGRESSION
# ============================================
ws5 = wb.create_sheet("5. Inferensi Manual")

ws5['A1'] = "SIMULASI FORWARD PASS LOGISTIC REGRESSION"
ws5['A1'].font = Font(bold=True, size=12)

ws5['A3'] = "Dokumen Uji: D1 ('web develop')"
ws5['A3'].font = Font(bold=True, color="FF0000")

# Vektor fitur dari Sheet 4
ws5['A5'] = "VEKTOR FITUR (x)"
ws5['A5'].font = Font(bold=True)

ws5['A6'] = "Fitur"
ws5['B6'] = "Nilai (x)"

for cell in ws5[6]:
    cell.fill = header_fill
    cell.font = header_font

# Ambil vektor D1 yang sudah dinormalisasi
vec_d1 = {
    "web": round(1.0000 / l2_norms[0], 4),
    "develop": round(1 * idf_values["develop"] / l2_norms[0], 4),
    "analyst": 0
}

features = ["web", "develop", "analyst"]
for i, feat in enumerate(features, start=7):
    ws5.cell(row=i, column=1, value=feat)
    ws5.cell(row=i, column=2, value=vec_d1[feat])

# Bobot koefisien (simulasi berdasarkan output model)
ws5['A11'] = "BOBOT KOEFISIEN MODEL (w)"
ws5['A11'].font = Font(bold=True)

ws5['A12'] = "Fitur"
ws5['B12'] = "Kelas Programmer"
ws5['C12'] = "Kelas Non-IT"
ws5['D12'] = "Kelas Data Analyst"
ws5['E12'] = "Kelas Wirausaha"

for cell in ws5[12]:
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Bobot koefisien (disesuaikan dengan output model)
weights = {
    "web": [0.7727, -0.5, -0.3, 0.2],
    "develop": [1.2192, -0.4, -0.2464, 0.1],
    "analyst": [-1.0188, -0.3, 2.8400, -0.6561]
}

for i, feat in enumerate(features, start=13):
    ws5.cell(row=i, column=1, value=feat)
    for j, w in enumerate(weights[feat], start=2):
        ws5.cell(row=i, column=j, value=w)

# Intersep (bias)
ws5['A17'] = "Intersep (b)"
ws5['B17'] = -0.5
ws5['C17'] = -0.895
ws5['D17'] = -1.2
ws5['E17'] = -0.45

for cell in ws5[17]:
    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Hitung skor linear z = w^T * x + b
ws5['A19'] = "SKOR LINEAR (z = w^T × x + b)"
ws5['A19'].font = Font(bold=True, size=12)

ws5['A20'] = "Kelas"
ws5['B20'] = "Skor (z)"

for cell in ws5[20]:
    cell.fill = header_fill
    cell.font = header_font

classes = ["Programmer", "Non-IT", "Data Analyst", "Wirausaha"]
z_scores = []

for i, cls in enumerate(classes, start=21):
    ws5.cell(row=i, column=1, value=cls)
    
    # Hitung z = sum(w * x) + b
    z = 0
    for feat in features:
        z += weights[feat][i-21] * vec_d1[feat]
    z += [-0.5, -0.895, -1.2, -0.45][i-21]
    
    z_scores.append(z)
    ws5.cell(row=i, column=2, value=round(z, 4))

# Hitung Softmax
ws5['A26'] = "PERHITUNGAN SOFTMAX (Probabilitas Posterior)"
ws5['A26'].font = Font(bold=True, size=12)

ws5['A27'] = "Kelas"
ws5['B27'] = "z"
ws5['C27'] = "e^z"
ws5['D27'] = "Probabilitas"

for cell in ws5[27]:
    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")

# Hitung e^z
exp_z = [math.exp(z) for z in z_scores]
sum_exp_z = sum(exp_z)

for i, (cls, z, e_z) in enumerate(zip(classes, z_scores, exp_z), start=28):
    ws5.cell(row=i, column=1, value=cls)
    ws5.cell(row=i, column=2, value=round(z, 4))
    ws5.cell(row=i, column=3, value=round(e_z, 4))
    
    prob = e_z / sum_exp_z
    ws5.cell(row=i, column=4, value=f"{prob:.4f} ({prob*100:.1f}%)")
    ws5.cell(row=i, column=4).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

ws5['A33'] = f"Total Penyebut (Σe^z): {sum_exp_z:.4f}"
ws5['A33'].font = Font(bold=True)

ws5['A35'] = "KESIMPULAN:"
ws5['A35'].font = Font(bold=True, color="FF0000", size=12)
ws5['A36'] = f"Probabilitas tertinggi: {max(exp_z)/sum_exp_z*100:.1f}% untuk kelas {classes[exp_z.index(max(exp_z))]}"
ws5['A37'] = "Karena probabilitas > 0.50 (threshold), status = auto_classified"

for col in ['A', 'B', 'C', 'D', 'E']:
    ws5.column_dimensions[col].width = 18

# Simpan file
filename = "Perhitungan_Manual_TFIDF_LogReg.xlsx"
wb.save(filename)

print(f"File Excel berhasil dibuat: {filename}")
print(f"\nSheet yang dibuat:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
